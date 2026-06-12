"""
高中网课销售追踪助手
工作台（今日计划）· 客户管理 · 跟进记录 · 推销助手 · 数据看板

运行：streamlit run sales_tracker.py
"""

import io
from datetime import date, datetime, timedelta

import streamlit as st

import tracker_db as db
import sales_engine as se
import knowledge_base as kb

st.set_page_config(
    page_title="高中网课销售追踪助手",
    page_icon="📚",
    layout="wide",
    initial_sidebar_state="expanded",
)

db.init_db()

# ─── CSS ──────────────────────────────────────────────────────────────────────

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+SC:wght@400;500;600;700&display=swap');
*, html, body, [class*="css"] { font-family: 'Noto Sans SC', -apple-system, sans-serif; }
.main .block-container { padding: 1.2rem 1.5rem 3rem; max-width: 1180px; }
.hero {
    background: linear-gradient(135deg, #0ea5e9 0%, #4f46e5 60%, #6d28d9 100%);
    border-radius: 18px; padding: 24px 32px; color: white; margin-bottom: 18px;
    box-shadow: 0 12px 40px rgba(79,70,229,0.30); position: relative; overflow: hidden;
}
.hero::after {
    content: '📚'; position: absolute; right: 28px; top: 50%;
    transform: translateY(-50%); font-size: 4.6rem; opacity: 0.13; pointer-events: none;
}
.hero h1 { margin: 0; font-size: 1.5rem; font-weight: 700; }
.hero p  { margin: 5px 0 0; opacity: 0.88; font-size: 0.9rem; }
.metric-card {
    background: white; border-radius: 14px; padding: 16px 18px;
    border: 1px solid #f1f5f9; box-shadow: 0 2px 12px rgba(0,0,0,.05); text-align: center;
}
.metric-num { font-size: 1.9rem; font-weight: 800; line-height: 1.2; }
.metric-label { font-size: .78rem; color: #94a3b8; margin-top: 2px; }
.task-row {
    display: flex; align-items: center; gap: 10px; flex-wrap: wrap;
    background: white; border-radius: 10px; padding: 10px 14px; margin-bottom: 8px;
    border: 1px solid #f1f5f9; box-shadow: 0 1px 6px rgba(0,0,0,.04);
}
.stage-pill {
    display: inline-block; padding: 2px 10px; border-radius: 99px;
    font-size: .74rem; font-weight: 600;
}
.sp-新线索 { background:#eff6ff; color:#1d4ed8; }
.sp-已联系 { background:#f5f3ff; color:#6d28d9; }
.sp-已试听 { background:#fffbeb; color:#b45309; }
.sp-谈单中 { background:#fff1f2; color:#be123c; }
.sp-已成交 { background:#f0fdf4; color:#15803d; }
.sp-已流失 { background:#f8fafc; color:#94a3b8; }
.script-block {
    background: #f8f9ff; border-left: 3px solid #818cf8;
    border-radius: 0 10px 10px 0; padding: 12px 16px; margin-bottom: 10px;
    font-size: .88rem; color: #334155; line-height: 1.8;
}
.plan-step {
    display: flex; gap: 12px; align-items: flex-start;
    padding: 9px 0; border-bottom: 1px solid #f1f5f9;
}
.plan-date {
    background: #4f46e5; color: white; border-radius: 6px; padding: 2px 9px;
    font-size: .76rem; font-weight: 700; flex-shrink: 0; margin-top: 1px; white-space: nowrap;
}
.section-title {
    font-size: 1.02rem; font-weight: 700; color: #1e293b;
    margin: 18px 0 12px; display: flex; align-items: center; gap: 8px;
}
.section-title::before {
    content: ''; display: block; width: 4px; height: 18px;
    background: linear-gradient(180deg, #0ea5e9, #6d28d9); border-radius: 2px;
}
#MainMenu, footer { visibility: hidden; }
@media (max-width: 640px) {
    .main .block-container { padding: 0.9rem; }
    .hero { padding: 16px 18px; }
    .hero h1 { font-size: 1.2rem; }
}
</style>
""", unsafe_allow_html=True)


# ─── 公共渲染 ─────────────────────────────────────────────────────────────────

def stage_pill(stage: str) -> str:
    return f'<span class="stage-pill sp-{stage}">{stage}</span>'


def score_color(score: int) -> str:
    if score >= 78: return "#059669"
    if score >= 62: return "#0284c7"
    if score >= 45: return "#d97706"
    return "#9ca3af"


def customer_line(c: dict, extra: str = "") -> str:
    nf = c.get("next_followup") or "未安排"
    sc = c.get("intent_score", 50)
    student = f"（{c['student_name']}）" if c.get("student_name") else ""
    return (
        f'<div class="task-row">'
        f'<b style="color:#1e293b;">{c["parent_name"]}</b>'
        f'<span style="color:#64748b;font-size:.82rem;">{student}{c["grade"]} · '
        f'{"、".join(c.get("subjects", [])[:3]) or "未指定科目"}</span>'
        f'{stage_pill(c["stage"])}'
        f'<span style="color:{score_color(sc)};font-weight:700;font-size:.84rem;">意向 {sc}</span>'
        f'<span style="color:#94a3b8;font-size:.78rem;margin-left:auto;">下次跟进：{nf}{extra}</span>'
        f'</div>'
    )


def customer_label(c: dict) -> str:
    student = f"/{c['student_name']}" if c.get("student_name") else ""
    return f"{c['parent_name']}{student} · {c['grade']} · {c['stage']}（意向{c['intent_score']}）"


def export_customers_excel(customers: list) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "客户列表"
    headers = ["家长称呼", "联系方式", "学生姓名", "年级", "科目", "预算(元/学期)",
               "成绩水平", "关注点", "意向信号", "来源", "阶段", "意向评分",
               "画像类型", "下次跟进", "备注", "创建时间"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="4F46E5")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = [12, 16, 10, 8, 18, 14, 10, 22, 22, 12, 8, 8, 10, 12, 30, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    for c in customers:
        ws.append([
            c["parent_name"], c.get("contact", ""), c.get("student_name", ""),
            c["grade"], "、".join(c.get("subjects", [])), c.get("budget", ""),
            c.get("performance", ""), "、".join(c.get("concerns", [])),
            "、".join(c.get("intent_signals", [])), c.get("source", ""),
            c["stage"], c.get("intent_score", ""), c.get("profile_type", ""),
            c.get("next_followup", ""), c.get("notes", ""), c.get("created_at", ""),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── 客户表单 ─────────────────────────────────────────────────────────────────

def customer_form(defaults: dict | None = None, key: str = "new"):
    d = defaults or {}
    c1, c2, c3 = st.columns(3)
    with c1:
        parent_name = st.text_input("家长称呼 *", value=d.get("parent_name", ""),
                                    placeholder="如：王女士", key=f"{key}_pn")
        student_name = st.text_input("学生姓名", value=d.get("student_name", ""), key=f"{key}_sn")
    with c2:
        contact = st.text_input("联系方式（微信/手机）", value=d.get("contact", ""), key=f"{key}_ct")
        grade = st.selectbox("年级 *", se.GRADE_OPTIONS,
                             index=se.GRADE_OPTIONS.index(d["grade"]) if d.get("grade") in se.GRADE_OPTIONS else 0,
                             key=f"{key}_gr")
    with c3:
        source = st.selectbox("客户来源", se.SOURCE_OPTIONS,
                              index=se.SOURCE_OPTIONS.index(d["source"]) if d.get("source") in se.SOURCE_OPTIONS else 0,
                              key=f"{key}_src")
        performance = st.selectbox("成绩水平", se.PERFORMANCE_OPTIONS,
                                   index=se.PERFORMANCE_OPTIONS.index(d["performance"]) if d.get("performance") in se.PERFORMANCE_OPTIONS else 2,
                                   key=f"{key}_pf")

    c4, c5 = st.columns(2)
    with c4:
        subjects = st.multiselect("咨询科目", se.SUBJECT_OPTIONS,
                                  default=[s for s in d.get("subjects", []) if s in se.SUBJECT_OPTIONS],
                                  key=f"{key}_sub")
        concerns = st.multiselect("家长关注点", se.CONCERN_OPTIONS,
                                  default=[s for s in d.get("concerns", []) if s in se.CONCERN_OPTIONS],
                                  key=f"{key}_cc")
    with c5:
        budget = st.slider("预算（元/学期）", 1000, 30000, d.get("budget", 6000), 500, key=f"{key}_bg")
        intent_signals = st.multiselect("意向信号", se.INTENT_OPTIONS,
                                        default=[s for s in d.get("intent_signals", []) if s in se.INTENT_OPTIONS],
                                        key=f"{key}_is")

    stage = st.selectbox("当前阶段", db.STAGES,
                         index=db.STAGES.index(d["stage"]) if d.get("stage") in db.STAGES else 0,
                         key=f"{key}_st")
    notes = st.text_area("备注（沟通要点/家庭背景等）", value=d.get("notes", ""),
                         height=70, key=f"{key}_nt")

    return {
        "parent_name": parent_name.strip(),
        "contact": contact.strip(),
        "student_name": student_name.strip(),
        "grade": grade,
        "subjects": subjects,
        "budget": budget,
        "performance": performance,
        "concerns": concerns,
        "intent_signals": intent_signals,
        "source": source,
        "stage": stage,
        "notes": notes.strip(),
        "next_followup": d.get("next_followup", ""),
    }


# ─── Tab 1：今日工作台 ────────────────────────────────────────────────────────

def render_dashboard():
    stats = db.get_funnel_stats()
    tasks = db.get_today_tasks()

    n_today = len(tasks["overdue"]) + len(tasks["today"])
    cols = st.columns(5)
    metrics = [
        ("跟进中客户", stats["active"], "#4f46e5"),
        ("今日待跟进", n_today, "#dc2626" if n_today else "#059669"),
        ("已逾期", len(tasks["overdue"]), "#dc2626" if tasks["overdue"] else "#94a3b8"),
        ("已成交", stats["won"], "#059669"),
        ("成交率", f'{stats["win_rate"]}%', "#0284c7"),
    ]
    for col, (label, num, color) in zip(cols, metrics):
        with col:
            st.markdown(
                f'<div class="metric-card"><div class="metric-num" style="color:{color};">{num}</div>'
                f'<div class="metric-label">{label}</div></div>',
                unsafe_allow_html=True,
            )

    st.markdown('<div class="section-title">📌 今日跟进计划</div>', unsafe_allow_html=True)
    if not tasks["overdue"] and not tasks["today"]:
        st.success("今天没有到期的跟进任务 🎉 可以去开发新线索，或处理下方长期未动客户")
    for c in tasks["overdue"]:
        days = (date.today() - date.fromisoformat(c["next_followup"])).days
        st.markdown(customer_line(c, f'　<b style="color:#dc2626;">已逾期 {days} 天</b>'),
                    unsafe_allow_html=True)
    for c in tasks["today"]:
        st.markdown(customer_line(c, "　<b style='color:#0284c7;'>今日到期</b>"), unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    with c1:
        st.markdown('<div class="section-title">🗓 未来3天</div>', unsafe_allow_html=True)
        if tasks["upcoming"]:
            for c in tasks["upcoming"]:
                st.markdown(customer_line(c), unsafe_allow_html=True)
        else:
            st.caption("未来3天暂无安排")
    with c2:
        st.markdown('<div class="section-title">🧊 超7天未跟进（流失预警）</div>', unsafe_allow_html=True)
        if tasks["stale"]:
            for c in tasks["stale"]:
                st.markdown(
                    customer_line(c, f'　<b style="color:#d97706;">闲置 {c.get("_idle_days", "?")} 天</b>'),
                    unsafe_allow_html=True)
        else:
            st.caption("没有被遗忘的客户，很好 👍")

    if tasks["no_plan"]:
        st.markdown('<div class="section-title">⚠️ 未安排下次跟进的活跃客户</div>', unsafe_allow_html=True)
        st.caption("以下客户没有设置下次跟进时间，建议尽快在「客户跟进」中安排：")
        for c in tasks["no_plan"]:
            st.markdown(customer_line(c), unsafe_allow_html=True)


# ─── Tab 2：客户管理 ──────────────────────────────────────────────────────────

def render_customers():
    with st.expander("➕ 新增客户", expanded=db.get_funnel_stats()["total"] == 0):
        data = customer_form(key="new")
        if st.button("保存新客户", type="primary", key="btn_add"):
            if not data["parent_name"]:
                st.error("请填写家长称呼")
            else:
                score = se.calc_intent_score(data)
                profile, _ = se.determine_profile(data)
                data["intent_score"] = score
                data["profile_type"] = profile
                cid = db.add_customer(data)
                st.success(f"已添加客户「{data['parent_name']}」（意向评分 {score}，画像：{profile}）")
                st.rerun()

    st.markdown('<div class="section-title">👥 客户列表</div>', unsafe_allow_html=True)
    f1, f2, f3, f4 = st.columns([1, 1, 1.6, 1])
    with f1:
        flt_stage = st.selectbox("阶段筛选", ["全部"] + db.STAGES, key="flt_stage")
    with f2:
        flt_grade = st.selectbox("年级筛选", ["全部"] + se.GRADE_OPTIONS, key="flt_grade")
    with f3:
        flt_kw = st.text_input("搜索（姓名/联系方式）", key="flt_kw")
    customers = db.list_customers(
        stage="" if flt_stage == "全部" else flt_stage,
        grade="" if flt_grade == "全部" else flt_grade,
        keyword=flt_kw.strip(),
    )
    with f4:
        st.markdown("<div style='height:1.75rem;'></div>", unsafe_allow_html=True)
        if customers:
            st.download_button(
                "📥 导出Excel", data=export_customers_excel(customers),
                file_name=f"客户列表_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    if not customers:
        st.info("暂无客户，先从上方「新增客户」开始录入吧")
        return

    for c in customers:
        student = f"（{c['student_name']}）" if c.get("student_name") else ""
        with st.expander(
            f"{c['parent_name']}{student} · {c['grade']} · {c['stage']} · 意向{c['intent_score']}"
            f" · {c.get('profile_type') or '未分析'}"
        ):
            edited = customer_form(defaults=c, key=f"edit_{c['id']}")
            b1, b2, b3 = st.columns([1, 1, 1])
            with b1:
                if st.button("💾 保存修改", key=f"save_{c['id']}"):
                    edited["intent_score"] = se.calc_intent_score(edited)
                    edited["profile_type"] = se.determine_profile(edited)[0]
                    db.update_customer(c["id"], edited)
                    st.success("已保存，意向评分与画像已重新计算")
                    st.rerun()
            with b2:
                if st.button("🗑 删除客户", key=f"del_{c['id']}"):
                    if st.session_state.get(f"confirm_del_{c['id']}"):
                        db.delete_customer(c["id"])
                        st.rerun()
                    else:
                        st.session_state[f"confirm_del_{c['id']}"] = True
                        st.warning("再点一次确认删除（含全部跟进记录）")
            with b3:
                st.caption(f"创建：{c['created_at'][:16]}　更新：{c['updated_at'][:16]}")


# ─── Tab 3：客户跟进 ──────────────────────────────────────────────────────────

def render_followups():
    customers = db.list_customers()
    if not customers:
        st.info("暂无客户，请先在「客户管理」中添加")
        return

    sel = st.selectbox("选择客户", customers, format_func=customer_label, key="fu_sel")
    c = sel

    st.markdown(customer_line(c), unsafe_allow_html=True)

    col_l, col_r = st.columns([1.1, 1])
    with col_l:
        st.markdown('<div class="section-title">📝 记录新跟进</div>', unsafe_allow_html=True)
        with st.form(f"fu_form_{c['id']}", clear_on_submit=True):
            f1, f2 = st.columns(2)
            with f1:
                fu_date = st.date_input("跟进日期", value=date.today())
                method = st.selectbox("跟进方式", db.FOLLOWUP_METHODS)
            with f2:
                result = st.selectbox("跟进结果", db.FOLLOWUP_RESULTS)
                next_date = st.date_input("下次跟进日期", value=date.today() + timedelta(days=3))
            content = st.text_area("沟通内容", placeholder="谈了什么、客户反馈、待办事项...", height=80)
            new_stage = st.selectbox("更新客户阶段", db.STAGES, index=db.STAGES.index(c["stage"]))
            if st.form_submit_button("保存跟进记录", type="primary", use_container_width=True):
                db.add_followup(c["id"], fu_date.isoformat(), method,
                                content.strip(), result, next_date.isoformat())
                if new_stage != c["stage"]:
                    db.set_stage(c["id"], new_stage)
                st.success("跟进已记录，下次跟进时间已更新到工作台")
                st.rerun()

    with col_r:
        st.markdown('<div class="section-title">🕘 历史跟进记录</div>', unsafe_allow_html=True)
        history = db.list_followups(c["id"])
        if not history:
            st.caption("还没有跟进记录")
        for h in history:
            next_str = f"　→ 下次：{h['next_date']}" if h.get("next_date") else ""
            st.markdown(f"""
            <div style="background:white;border-radius:10px;padding:10px 14px;margin-bottom:8px;
                        border:1px solid #f1f5f9;border-left:3px solid #818cf8;">
              <div style="font-size:.78rem;color:#94a3b8;">
                {h['fu_date']} · {h['method']} ·
                <b style="color:#4f46e5;">{h['result']}</b>{next_str}
              </div>
              <div style="font-size:.86rem;color:#334155;margin-top:4px;line-height:1.6;">
                {h['content'] or '（未填写内容）'}
              </div>
            </div>""", unsafe_allow_html=True)


# ─── Tab 4：推销助手 ──────────────────────────────────────────────────────────

# 推销助手画像 → 知识库画像名映射（知识库使用 K12 通用画像名）
_KB_CONCERN_MAP = {
    "高考冲刺": "升学", "薄弱科目补差": "补差", "提分": "提分",
    "自主招生/强基": "冲刺名校", "学习方法": "培养兴趣",
}


def render_assistant():
    customers = db.list_customers()
    if not customers:
        st.info("暂无客户，请先在「客户管理」中添加")
        return

    sel = st.selectbox("选择客户", customers, format_func=customer_label, key="as_sel")
    c = sel
    analysis = se.analyze_customer(c)
    score = analysis["intent_score"]
    color = score_color(score)

    # 概览卡
    c1, c2 = st.columns([1, 2.2])
    with c1:
        st.markdown(f"""
        <div class="metric-card" style="height:100%;">
          <div class="metric-label">购买意向评分</div>
          <div class="metric-num" style="font-size:2.8rem;color:{color};">{score}</div>
          <div style="background:#e5e7eb;border-radius:99px;height:8px;margin:8px 0;overflow:hidden;">
            <div style="width:{score}%;height:100%;background:{color};border-radius:99px;"></div>
          </div>
          <div style="font-size:.82rem;color:{color};font-weight:600;">⏱ {analysis['closing_cycle']}</div>
        </div>""", unsafe_allow_html=True)
    with c2:
        st.markdown(f"""
        <div style="background:white;border-radius:14px;padding:18px 20px;height:100%;
                    border:1px solid #f1f5f9;box-shadow:0 2px 12px rgba(0,0,0,.05);">
          <span style="background:linear-gradient(135deg,#0ea5e9,#6d28d9);color:white;
                       padding:4px 16px;border-radius:20px;font-weight:700;font-size:.95rem;">
            {analysis['profile_type']}
          </span>
          <div style="font-size:.86rem;color:#475569;line-height:1.75;margin-top:10px;">
            {analysis['profile_description']}
          </div>
        </div>""", unsafe_allow_html=True)

    # 阶段行动建议
    sa = analysis["stage_actions"]
    st.markdown(f'<div class="section-title">🎯 当前阶段「{c["stage"]}」行动建议 · 目标：{sa["goal"]}</div>',
                unsafe_allow_html=True)
    rows = "".join(
        f'<div class="plan-step"><div class="plan-date">{i}</div>'
        f'<div style="font-size:.87rem;color:#475569;line-height:1.6;">{a}</div></div>'
        for i, a in enumerate(sa["actions"], 1)
    )
    st.markdown(f'<div style="background:white;border-radius:12px;padding:8px 16px;'
                f'border:1px solid #f1f5f9;">{rows}</div>', unsafe_allow_html=True)

    # 话术
    st.markdown('<div class="section-title">💬 推荐话术（针对该画像 · 网课场景）</div>', unsafe_allow_html=True)
    for i, script in enumerate(analysis["scripts"], 1):
        st.markdown(f"""
        <div class="script-block">
          <div style="font-size:.73rem;color:#6366f1;font-weight:700;margin-bottom:5px;">话术 {i}</div>
          {script}
        </div>""", unsafe_allow_html=True)

    # 异议应对
    st.markdown('<div class="section-title">🛡 网课常见异议应对</div>', unsafe_allow_html=True)
    objection = st.selectbox("客户提出了什么异议？", list(se.OBJECTION_HANDLERS.keys()), key="obj_sel")
    st.markdown(f'<div class="script-block" style="border-left-color:#f59e0b;background:#fffbeb;">'
                f'{se.OBJECTION_HANDLERS[objection]}</div>', unsafe_allow_html=True)

    # 卖点速查
    with st.expander("📦 网课核心卖点速查"):
        for p in se.COURSE_SELLING_POINTS:
            st.markdown(f"- {p}")

    # SPIN / 影响力策略依据
    with st.expander("📚 策略依据（SPIN 提问 + 影响力原则）"):
        kb_data = {
            "concerns": [_KB_CONCERN_MAP.get(x, x) for x in c.get("concerns", [])],
            "subjects": c.get("subjects", []),
        }
        refs = kb.get_kb_references(analysis["profile_type"], kb_data)
        spin = refs["spin_focus"]
        st.markdown(
            f"**SPIN 提问侧重**：S情境 `{spin['S']}` · P问题 `{spin['P']}` · "
            f"I暗示 `{spin['I']}` · N价值 `{spin['N']}`\n\n> {spin.get('note', '')}"
        )
        st.markdown("**优先影响力原则**：")
        for p in refs["influence_principles"]:
            st.markdown(f"- **{p['name']}**（{p['timing']}）：{' / '.join(p['actions'])}")
        if refs["spin_question_examples"]:
            st.markdown("**参考提问**：")
            for q in refs["spin_question_examples"]:
                st.markdown(f"- ❓ {q}")

    # 跟进计划
    st.markdown('<div class="section-title">🗓 智能跟进计划</div>', unsafe_allow_html=True)
    plan = analysis["plan"]
    rows = "".join(
        f'<div class="plan-step"><div class="plan-date">{p["date"][5:]} (D+{p["day_offset"]})</div>'
        f'<div style="font-size:.87rem;color:#475569;line-height:1.6;">{p["action"]}</div></div>'
        for p in plan
    )
    st.markdown(f'<div style="background:white;border-radius:12px;padding:8px 16px;'
                f'border:1px solid #f1f5f9;">{rows}</div>', unsafe_allow_html=True)

    p1, p2 = st.columns([1, 2])
    with p1:
        if st.button("📌 采用此计划（写入下次跟进时间）", key="adopt_plan", type="primary"):
            first_future = next(
                (p["date"] for p in plan if p["date"] > date.today().isoformat()),
                plan[-1]["date"],
            )
            db.set_next_followup(c["id"], first_future)
            db.update_customer(c["id"], {**c, "intent_score": score,
                                         "profile_type": analysis["profile_type"],
                                         "next_followup": first_future})
            st.success(f"已将下次跟进时间设为 {first_future}，可在工作台查看")
    with p2:
        st.caption("采用后，计划中最近的一个未来日期会写入该客户的「下次跟进」字段，并同步更新画像与评分。")


# ─── Tab 5：数据看板 ──────────────────────────────────────────────────────────

def render_stats():
    stats = db.get_funnel_stats()
    if stats["total"] == 0:
        st.info("暂无数据，添加客户后这里会显示销售漏斗与转化分析")
        return

    st.markdown('<div class="section-title">🔻 销售漏斗</div>', unsafe_allow_html=True)
    max_n = max(stats["by_stage"].values()) or 1
    funnel_colors = {
        "新线索": "#3b82f6", "已联系": "#8b5cf6", "已试听": "#f59e0b",
        "谈单中": "#f43f5e", "已成交": "#22c55e", "已流失": "#94a3b8",
    }
    bars = ""
    for stage in db.STAGES:
        n = stats["by_stage"].get(stage, 0)
        w = max(6, round(n / max_n * 100))
        color = funnel_colors[stage]
        bars += (
            f'<div style="display:flex;align-items:center;gap:10px;margin-bottom:8px;">'
            f'<div style="width:64px;font-size:.84rem;color:#475569;font-weight:600;flex-shrink:0;">{stage}</div>'
            f'<div style="flex:1;background:#f1f5f9;border-radius:8px;overflow:hidden;">'
            f'<div style="width:{w}%;background:{color};border-radius:8px;padding:5px 10px;'
            f'color:white;font-size:.82rem;font-weight:700;min-width:32px;">{n}</div></div></div>'
        )
    st.markdown(f'<div style="background:white;border-radius:14px;padding:18px 20px;'
                f'border:1px solid #f1f5f9;">{bars}</div>', unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    with c1:
        st.markdown('<div class="section-title">📈 转化概览</div>', unsafe_allow_html=True)
        bs = stats["by_stage"]
        contacted = stats["total"] - bs.get("新线索", 0)
        tried = bs.get("已试听", 0) + bs.get("谈单中", 0) + bs.get("已成交", 0)
        rows = [
            ("客户总数", stats["total"], ""),
            ("已触达", contacted, f"{round(contacted / stats['total'] * 100)}%" if stats["total"] else "-"),
            ("进入试听及以后", tried, f"{round(tried / stats['total'] * 100)}%" if stats["total"] else "-"),
            ("已成交", stats["won"], f"{round(stats['won'] / stats['total'] * 100)}%" if stats["total"] else "-"),
            ("成交率（成交/已关单）", f"{stats['win_rate']}%", ""),
        ]
        html = ""
        for lbl, val, pct in rows:
            pct_html = f' <span style="color:#94a3b8;font-weight:400;">({pct})</span>' if pct else ""
            html += (
                f'<div style="display:flex;justify-content:space-between;padding:8px 0;'
                f'border-bottom:1px solid #f1f5f9;font-size:.88rem;">'
                f'<span style="color:#475569;">{lbl}</span>'
                f'<span style="font-weight:700;color:#1e293b;">{val}{pct_html}</span></div>'
            )
        st.markdown(f'<div style="background:white;border-radius:14px;padding:12px 20px;'
                    f'border:1px solid #f1f5f9;">{html}</div>', unsafe_allow_html=True)

    with c2:
        st.markdown('<div class="section-title">📣 来源渠道分析</div>', unsafe_allow_html=True)
        src = db.get_source_stats()
        rows_html = ""
        for name, s in sorted(src.items(), key=lambda kv: -kv[1]["total"]):
            rate = f"{round(s['won'] / s['total'] * 100)}%" if s["total"] else "-"
            rows_html += (
                f'<div style="display:flex;justify-content:space-between;padding:8px 0;'
                f'border-bottom:1px solid #f1f5f9;font-size:.88rem;">'
                f'<span style="color:#475569;">{name}</span>'
                f'<span style="color:#1e293b;">共 <b>{s["total"]}</b> · 成交 <b>{s["won"]}</b>'
                f' · 转化 <b style="color:#059669;">{rate}</b></span></div>'
            )
        st.markdown(f'<div style="background:white;border-radius:14px;padding:12px 20px;'
                    f'border:1px solid #f1f5f9;">{rows_html}</div>', unsafe_allow_html=True)

    # 高意向客户排行
    st.markdown('<div class="section-title">🔥 高意向客户 TOP10（未成交）</div>', unsafe_allow_html=True)
    hot = [c for c in db.list_customers() if c["stage"] in db.ACTIVE_STAGES][:10]
    if hot:
        for c in hot:
            st.markdown(customer_line(c), unsafe_allow_html=True)
    else:
        st.caption("暂无活跃客户")


# ─── 主入口 ───────────────────────────────────────────────────────────────────

def main():
    with st.sidebar:
        st.markdown("""
        <div style="text-align:center;padding:10px 0 16px;">
          <div style="font-size:2.4rem;">📚</div>
          <div style="font-weight:700;font-size:1rem;color:#1e293b;margin-top:6px;">高中网课销售追踪助手</div>
          <div style="font-size:.78rem;color:#94a3b8;margin-top:3px;">客户追踪 · 话术推荐 · 计划制定</div>
        </div>""", unsafe_allow_html=True)
        st.divider()
        stats = db.get_funnel_stats()
        tasks = db.get_today_tasks()
        n_today = len(tasks["overdue"]) + len(tasks["today"])
        st.markdown(f"""
        <div style="font-size:.85rem;color:#475569;line-height:2.1;">
          👥 跟进中客户：<b>{stats['active']}</b><br>
          📌 今日待跟进：<b style="color:{'#dc2626' if n_today else '#059669'};">{n_today}</b><br>
          ✅ 累计成交：<b style="color:#059669;">{stats['won']}</b>
        </div>""", unsafe_allow_html=True)
        st.divider()
        st.markdown("""
        <div style="font-size:.8rem;color:#64748b;line-height:1.9;">
          <b>📌 使用流程</b><br>
          ① 客户管理：录入新线索<br>
          ② 推销助手：看画像/话术/计划<br>
          ③ 客户跟进：记录每次沟通<br>
          ④ 工作台：每天按计划跟进<br>
          ⑤ 数据看板：复盘转化效果
        </div>""", unsafe_allow_html=True)
        st.divider()
        st.markdown("""
        <div style="font-size:.72rem;color:#cbd5e1;text-align:center;line-height:1.8;">
          数据保存在本地 sales_tracker.db<br>不上传任何客户信息
        </div>""", unsafe_allow_html=True)

    st.markdown(f"""
    <div class="hero">
      <h1>📚 高中网课销售追踪助手</h1>
      <p>{date.today().strftime('%Y年%m月%d日')} · 客户画像 · 跟进追踪 · 智能计划 · 提升成单率</p>
    </div>""", unsafe_allow_html=True)

    tab1, tab2, tab3, tab4, tab5 = st.tabs(
        ["📋 今日工作台", "👥 客户管理", "📝 客户跟进", "💡 推销助手", "📈 数据看板"]
    )
    with tab1:
        render_dashboard()
    with tab2:
        render_customers()
    with tab3:
        render_followups()
    with tab4:
        render_assistant()
    with tab5:
        render_stats()


if __name__ == "__main__":
    main()
