import streamlit as st
import json
import re
import io
from pathlib import Path
from datetime import datetime
import knowledge_base as kb

st.set_page_config(
    page_title="K12教培客户画像分析器",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+SC:wght@400;500;600;700&display=swap');

*, html, body, [class*="css"] {
    font-family: 'Noto Sans SC', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
}
.main .block-container {
    padding: 1.5rem 1.5rem 3rem;
    max-width: 1080px;
}
.hero {
    background: linear-gradient(135deg, #4f46e5 0%, #7c3aed 60%, #6d28d9 100%);
    border-radius: 18px;
    padding: 28px 36px;
    color: white;
    margin-bottom: 24px;
    box-shadow: 0 12px 40px rgba(79,70,229,0.32);
    position: relative;
    overflow: hidden;
}
.hero::after {
    content: '🎓';
    position: absolute;
    right: 28px; top: 50%;
    transform: translateY(-50%);
    font-size: 5rem;
    opacity: 0.12;
    pointer-events: none;
}
.hero h1 { margin: 0; font-size: 1.65rem; font-weight: 700; }
.hero p  { margin: 6px 0 0; opacity: 0.88; font-size: 0.93rem; }

.card {
    background: white;
    border-radius: 14px;
    padding: 22px 24px;
    box-shadow: 0 2px 16px rgba(0,0,0,0.055);
    margin-bottom: 18px;
    border: 1px solid #f1f5f9;
}
.card-title {
    font-size: 0.95rem; font-weight: 700; color: #1e293b;
    margin-bottom: 16px;
    padding-bottom: 10px;
    border-bottom: 1px solid #f1f5f9;
}
.result-section {
    font-size: 1.05rem; font-weight: 700; color: #1e293b;
    margin: 22px 0 14px;
    display: flex; align-items: center; gap: 8px;
}
.result-section::before {
    content: '';
    display: block;
    width: 4px; height: 20px;
    background: linear-gradient(180deg, #4f46e5, #7c3aed);
    border-radius: 2px;
    flex-shrink: 0;
}
.profile-pill {
    display: inline-block;
    background: linear-gradient(135deg, #4f46e5, #7c3aed);
    color: white;
    padding: 5px 18px;
    border-radius: 20px;
    font-size: 1.05rem;
    font-weight: 700;
    margin-bottom: 10px;
}
.tag {
    display: inline-block;
    padding: 3px 10px;
    border-radius: 6px;
    font-size: 0.82rem;
    font-weight: 500;
    margin: 2px;
}
.tag-blue   { background: #eff6ff; color: #1d4ed8; }
.tag-purple { background: #f5f3ff; color: #6d28d9; }
.tag-green  { background: #f0fdf4; color: #15803d; }
.script-block {
    background: #f8f9ff;
    border-left: 3px solid #818cf8;
    border-radius: 0 10px 10px 0;
    padding: 13px 17px;
    margin-bottom: 10px;
    font-size: 0.88rem;
    color: #334155;
    line-height: 1.8;
}
.info-box {
    border-radius: 10px;
    padding: 14px 16px;
    font-size: 0.88rem;
    line-height: 1.65;
    margin-bottom: 10px;
}
.ib-blue   { background: #eff6ff; border-left: 3px solid #3b82f6; color: #1e40af; }
.ib-purple { background: #f5f3ff; border-left: 3px solid #8b5cf6; color: #5b21b6; }
.tip-row {
    display: flex; align-items: flex-start; gap: 10px;
    padding: 9px 0; border-bottom: 1px solid #f1f5f9;
}
.tip-num {
    background: #4f46e5; color: white;
    border-radius: 4px; padding: 1px 7px;
    font-size: 0.74rem; font-weight: 700;
    flex-shrink: 0; margin-top: 2px;
}
.stButton > button {
    width: 100%;
    background: linear-gradient(135deg, #4f46e5, #7c3aed) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    padding: 14px !important;
    font-size: 1rem !important;
    font-weight: 600 !important;
    box-shadow: 0 4px 20px rgba(79,70,229,0.35) !important;
    transition: all .2s !important;
}
.stButton > button:hover {
    box-shadow: 0 6px 26px rgba(79,70,229,0.52) !important;
    transform: translateY(-1px) !important;
}
#MainMenu, footer, header { visibility: hidden; }
@media (max-width: 640px) {
    .main .block-container { padding: 0.9rem; }
    .hero { padding: 18px 20px; }
    .hero h1 { font-size: 1.25rem; }
}
</style>
""", unsafe_allow_html=True)

# ─── Config (API key persistence) ────────────────────────────────────────────

_CONFIG_PATH = Path(__file__).parent / "config.json"

def load_config() -> dict:
    if _CONFIG_PATH.exists():
        try:
            return json.loads(_CONFIG_PATH.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}

def save_config(config: dict):
    _CONFIG_PATH.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")


# ─── Excel Export ─────────────────────────────────────────────────────────────

_HISTORY_PATH = Path(__file__).parent / "analysis_history.xlsx"
_HISTORY_HEADERS = [
    "分析时间", "家长称呼", "孩子年级", "咨询科目", "家庭预算(元/学期)",
    "成绩水平", "关注点", "意向信号", "画像类型", "购买意向评分",
    "预计成交周期", "分析方式", "沟通备注",
]

def append_to_history(data: dict, result: dict):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return
    if _HISTORY_PATH.exists():
        wb = openpyxl.load_workbook(_HISTORY_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "分析历史"
        ws.append(_HISTORY_HEADERS)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor="4F46E5")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22
        col_widths = [16, 10, 10, 16, 16, 10, 20, 24, 12, 10, 16, 24, 36]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    row = [
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        data.get("parent_name", ""),
        data.get("grade", ""),
        "、".join(data.get("subjects", [])),
        data.get("budget", ""),
        data.get("performance", ""),
        "、".join(data.get("concerns", [])),
        "、".join(data.get("intent_signals", [])),
        result.get("profile_type", ""),
        result.get("purchase_intent_score", ""),
        result.get("closing_cycle", ""),
        result.get("analysis_method", ""),
        data.get("notes", ""),
    ]
    ws.append(row)
    wb.save(_HISTORY_PATH)


def generate_report_bytes(data: dict, result: dict) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return b""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "客户分析报告"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 58

    def _title(r, text):
        ws.merge_cells(f"A{r}:B{r}")
        c = ws.cell(r, 1, text)
        c.font = Font(bold=True, color="FFFFFF", size=15)
        c.fill = PatternFill("solid", fgColor="4F46E5")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 34
        return r + 1

    def _section(r, text):
        ws.merge_cells(f"A{r}:B{r}")
        c = ws.cell(r, 1, text)
        c.font = Font(bold=True, color="4F46E5", size=11)
        c.fill = PatternFill("solid", fgColor="EEF2FF")
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 22
        return r + 1

    def _row(r, label, value, h=18):
        ca = ws.cell(r, 1, label)
        ca.font = Font(bold=True, color="334155", size=10)
        ca.fill = PatternFill("solid", fgColor="F8FAFC")
        ca.alignment = Alignment(vertical="top", wrap_text=True)
        cb = ws.cell(r, 2, str(value) if value is not None else "")
        cb.font = Font(color="1E293B", size=10)
        cb.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = h
        return r + 1

    def _blank(r):
        ws.row_dimensions[r].height = 7
        return r + 1

    r = 1
    r = _title(r, "K12教培客户画像分析报告")
    r = _row(r, "生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    r = _row(r, "分析方式", result.get("analysis_method", ""))
    r = _blank(r)

    r = _section(r, "📋  基本信息")
    r = _row(r, "家长称呼", data.get("parent_name", "") or "未填写")
    r = _row(r, "孩子年级", data.get("grade", ""))
    r = _row(r, "咨询科目", "、".join(data.get("subjects", [])) or "未选择")
    r = _row(r, "家庭预算", f"¥ {data.get('budget', 0):,} 元/学期")
    r = _row(r, "成绩水平", data.get("performance", ""))
    r = _row(r, "关注点", "、".join(data.get("concerns", [])) or "未选择")
    r = _row(r, "意向信号", "、".join(data.get("intent_signals", [])) or "未选择")
    if data.get("notes"):
        r = _row(r, "沟通备注", data["notes"], h=50)
    r = _blank(r)

    r = _section(r, "🎯  画像分析")
    r = _row(r, "画像类型", result.get("profile_type", ""))
    r = _row(r, "购买意向评分", f"{result.get('purchase_intent_score', 0)} / 100")
    r = _row(r, "预计成交周期", result.get("closing_cycle", ""))
    r = _row(r, "画像描述", result.get("profile_description", ""), h=60)
    r = _row(r, "核心诉求", "  |  ".join(result.get("core_needs", [])))
    r = _blank(r)

    r = _section(r, "✅  机会点")
    for opp in result.get("opportunity_points", []):
        r = _row(r, "▶", opp)
    r = _blank(r)

    r = _section(r, "⚠️  风险点")
    for risk in result.get("risk_factors", []):
        r = _row(r, "▶", risk)
    r = _blank(r)

    r = _section(r, "📋  跟进策略")
    strategy = result.get("follow_up_strategy", {})
    r = _row(r, "跟进频率", strategy.get("follow_up_frequency", ""), h=30)
    r = _row(r, "促单时机", strategy.get("closing_timing", ""), h=40)
    r = _blank(r)
    for i, script in enumerate(strategy.get("script_templates", []), 1):
        r = _row(r, f"话术 {i}", script, h=72)
    r = _blank(r)
    for i, tip in enumerate(strategy.get("communication_tips", []), 1):
        r = _row(r, f"注意事项 {i}", tip, h=26)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Constants ────────────────────────────────────────────────────────────────

GRADE_OPTIONS = [
    "小学一年级", "小学二年级", "小学三年级",
    "小学四年级", "小学五年级", "小学六年级",
    "初一", "初二", "初三",
    "高一", "高二", "高三",
]

SUBJECT_OPTIONS = ["语文", "数学", "英语", "物理", "化学", "生物", "全科辅导"]

PERFORMANCE_OPTIONS = ["班级前10%", "中等偏上", "中等", "中等偏下", "较差"]

CONCERN_OPTIONS = ["提分", "升学", "培养兴趣", "补差", "冲刺名校", "作业辅导"]

INTENT_OPTIONS = [
    "主动咨询", "询问价格", "询问师资", "要求试听",
    "犹豫不决", "对比竞品", "已在其他机构上课",
]

# ─── Rule-based Engine ────────────────────────────────────────────────────────

def _calc_score(data: dict) -> int:
    score = 42
    signal_delta = {
        "主动咨询": 15, "要求试听": 20, "询问价格": 8, "询问师资": 7,
        "犹豫不决": -12, "对比竞品": -8, "已在其他机构上课": -15,
    }
    for s in data["intent_signals"]:
        score += signal_delta.get(s, 0)

    perf_delta = {"较差": 12, "中等偏下": 9, "中等": 2, "中等偏上": 0, "班级前10%": 3}
    score += perf_delta.get(data["performance"], 0)

    if "冲刺名校" in data["concerns"]: score += 10
    if "升学"   in data["concerns"]: score += 7
    if "补差"   in data["concerns"]: score += 7

    b = data["budget"]
    if b >= 20000: score += 10
    elif b >= 10000: score += 5
    elif b <= 3000: score -= 8

    grade = data["grade"]
    if grade in ("初三", "高三"): score += 12
    elif grade in ("初二", "高二"): score += 8
    elif grade in ("初一", "高一"): score += 4

    return max(5, min(97, score))


def _determine_profile(data: dict) -> tuple:
    concerns = set(data["concerns"])
    signals  = set(data["intent_signals"])
    budget   = data["budget"]
    perf     = data["performance"]

    pts = {"名校执念型": 0, "焦虑型": 0, "价格敏感型": 0,
           "理性分析型": 0, "素质培养型": 0, "佛系型": 0}

    if "冲刺名校" in concerns: pts["名校执念型"] += 30
    if "升学"   in concerns: pts["名校执念型"] += 15
    if budget >= 15000:       pts["名校执念型"] += 10

    if perf in ("中等偏下", "较差"): pts["焦虑型"] += 25
    if "补差" in concerns:          pts["焦虑型"] += 20
    if "提分" in concerns:          pts["焦虑型"] += 10
    if data["grade"] in ("初三", "高三"): pts["焦虑型"] += 15

    if "询问价格" in signals: pts["价格敏感型"] += 25
    if budget <= 5000:        pts["价格敏感型"] += 20
    if "对比竞品" in signals: pts["价格敏感型"] += 10

    if "询问师资" in signals:  pts["理性分析型"] += 20
    if len(signals) >= 3:      pts["理性分析型"] += 15
    if "对比竞品" in signals:  pts["理性分析型"] += 10
    if budget >= 10000:        pts["理性分析型"] += 5

    if "培养兴趣" in concerns:      pts["素质培养型"] += 30
    if perf in ("班级前10%", "中等偏上"): pts["素质培养型"] += 10

    if "犹豫不决" in signals: pts["佛系型"] += 25
    if len(signals) <= 1:     pts["佛系型"] += 15

    profile = max(pts, key=pts.get)
    descs = {
        "名校执念型": "对升学和知名院校有强烈执念，教育投入意愿高。决策时高度关注机构的升学成绩和师资背景，对价格不敏感，更看重结果保障和口碑背书。行动力较强，但也会货比三家顶尖机构。",
        "焦虑型":    "孩子成绩不理想使家长情绪焦虑明显，需求迫切。容易被紧迫感和逆袭案例打动，决策速度快，但情绪波动时也可能反复。需要专业引导和情感共情，同时给予清晰的提升方案。",
        "价格敏感型": "把性价比放在首位，倾向于多方比价和询问优惠。对折扣和活动敏感，建立信任后愿意付费，但前期需要充分的价值铺垫。适合通过体验课和分期方案降低决策门槛。",
        "理性分析型": "思维严谨，注重收集信息和多维对比。需要详细的课程方案、师资介绍和效果数据，不喜欢被催促。决策周期较长，但一旦认可则较为稳定，后续转介绍率高。",
        "素质培养型": "注重孩子全面发展和兴趣养成，对分数执念不深。需要强调教学理念、思维培养和学习习惯建设，展示孩子综合成长而非单纯分数提升。沟通要轻松不功利。",
        "佛系型":    "需求感不强烈，购买意愿偏低，可能是被动了解或随意咨询。缺乏明显紧迫感，短期难以成交。需要长期培育信任关系，等待时机，一旦出现需求触发点要迅速跟进。",
    }
    return profile, descs[profile]


def _get_strategy(profile: str, data: dict) -> dict:
    name  = data["parent_name"] or "家长"
    grade = data["grade"]
    perf  = data["performance"]
    per_c = max(50, round(data["budget"] / 40))

    templates = {
        "名校执念型": {
            "scripts": [
                f"「{name}您好，我看您非常重视孩子的升学。我们机构今年有5名同学成功考入市重点，我把其中一位{grade}学员的备考规划发给您参考，看看是否适合您孩子？」",
                f"「针对冲刺重点校，我们有专项班和专属升学顾问，从{grade}开始规划正是最佳时机。上周刚有一位家长加入，效果很好——要不要帮您约一次免费的升学规划诊断？」",
                f"「您关注的几所重点学校，我们都深度研究过它们的招生特点和考察方向，可以为您孩子定制专属方案。周末有个公开课，名额很紧张，我帮您预留一个？」",
            ],
            "follow_up_frequency": "每3-4天跟进一次，每次携带新内容（升学案例/名师资料/活动预告）",
            "closing_timing": "成绩单发放后；距离重要考试还有3-4个月时；机构限额名班即将开课前",
            "tips": ["重点展示往届升学成绩和真实学员案例", "提供针对性的升学路径规划图", "强调名额稀缺性制造适当紧迫感", "避免主动谈价格，让家长先认可价值"],
        },
        "焦虑型": {
            "scripts": [
                f"「{name}，我完全理解您的担心。孩子现在{perf}，这种情况我们处理过很多次，都成功帮孩子实现了逆转。最近有个和您孩子情况相似的{grade}学员，两个月提了18分，想听听他的经历吗？」",
                f"「建议先让孩子来上一节免费诊断课，我们老师经验丰富，一课就能找到问题根源，不占太多时间。这周哪天方便带孩子来？」",
                f"「越早开始越好，{grade}的孩子学习习惯还有很大的调整空间，晚一个月差距就会拉大。我刚查了一下，本周还有两个名额，您要现在帮孩子预留一下吗？」",
            ],
            "follow_up_frequency": "每2-3天跟进一次，及时回应消息，保持热度不断线",
            "closing_timing": "考试或作业成绩出来后（家长最焦虑时）；学期初新阶段开始时；孩子某次重要考试失利后",
            "tips": ["先用同理心安抚情绪，再给专业解决方案", "多分享成功逆袭案例，增强家长信心", "适当强调时间紧迫性，但不要过度施压", "避免让家长觉得是在批评孩子"],
        },
        "价格敏感型": {
            "scripts": [
                f"「{name}，我帮您算一笔：我们一学期课程平均每节才{per_c}元，比同类机构低15-20%，而且有阶段性提分承诺。这个性价比在市面上真的难找。」",
                f"「正好赶上我们开学季特惠，本月报名85折，还赠价值500元的学习礼包，活动到月底截止。您要不先锁定这个价格，后面价格会恢复正常的？」",
                f"「我理解您想先看看效果再决定。我们有一个月体验班，价格更灵活，满意了再续报，不满意全退——这样您的风险降到最低，您看怎么样？」",
            ],
            "follow_up_frequency": "每周跟进一次，促销活动前3天密集接触",
            "closing_timing": "限时优惠截止前；开学季/寒暑假前；机构搞活动时提前预告制造期待",
            "tips": ["用清晰数字展示性价比，不要虚夸", "主动提供分期付款或体验班选项", "节点性优惠要提前告知，制造期待", "不要单纯让价，先把价值说清楚再谈价格"],
        },
        "理性分析型": {
            "scripts": [
                f"「{name}，根据孩子的情况，我给您整理了一份详细的课程方案和近期学员成绩对比数据，以及我们师资介绍，发到您微信方便吗？有任何疑问随时沟通。」",
                f"「我们的教学体系分三个阶段：第一阶段诊断薄弱点、第二阶段专项强化、第三阶段综合提升，每阶段都有测评数据，家长随时可以看到进展。」",
                f"「如果您还在对比其他机构，我可以提供一份客观的对比清单——从师资背景、课程体系、服务保障三个维度分析，帮您做更准确的判断，完全不带倾向性。」",
            ],
            "follow_up_frequency": "每周一次，发送有价值的内容（学习资料/机构介绍/行业报告），不催促",
            "closing_timing": "提供完整材料后等待自然决策周期；家长主动回来追问细节时趁热打铁",
            "tips": ["提供详细课程大纲和师资背景材料", "用数据和事实说话，不夸大效果", "给足思考时间，切勿频繁催促", "可主动提供竞品对比，展示自信"],
        },
        "素质培养型": {
            "scripts": [
                f"「{name}，我们特别注重学习习惯和思维方式的培养，不只追分数。很多家长反馈孩子来了之后，不仅成绩稳住了，自主学习的意愿和对{'/'.join(data['subjects'][:2]) if data['subjects'] else '学科'}的兴趣也明显增强了。」",
                f"「我们老师在课堂上强调启发式教学，让孩子真正理解知识而不是死记硬背。我可以安排一次公开体验课，您亲眼看看我们的教学方式，比我说什么都直观。」",
                f"「除课程外，我们每月会给家长发送三维成长报告——学习态度、思维方式、知识掌握各一个维度，帮助家长全面了解孩子的真实成长状态。」",
            ],
            "follow_up_frequency": "每10天左右温和接触，分享教育理念类内容而非促销信息",
            "closing_timing": "孩子对某个学科产生明显兴趣或困惑时；学期末做下学期规划时",
            "tips": ["强调教学理念和孩子综合成长，不只谈提分", "展示孩子能力提升和习惯养成的真实案例", "保持轻松朋友式沟通，尊重家长理念", "避免用焦虑话术，这类家长对焦虑营销反感"],
        },
        "佛系型": {
            "scripts": [
                f"「{name}，完全没有关系，您慢慢考虑。我先把我们的课程资料发给您，不用有任何压力，有问题随时找我聊。」",
                f"「我最近整理了一份{grade}阶段学习规划的干货资料，觉得对您孩子很有参考价值，纯分享不推销，发给您看看？」",
                f"「下周我们有一个免费公开讲座，专门讲{grade}阶段的高效学习方法，很多家长听完觉得很实用，有兴趣带孩子来感受一下吗？」",
            ],
            "follow_up_frequency": "每2周轻触一次，以输出价值为主，保持存在感但不给压力",
            "closing_timing": "孩子成绩出现明显波动时；升学关键节点临近时（如小升初/中考/高考倒计时）；朋友圈晒孩子成绩时",
            "tips": ["保持低压力的朋友式沟通方式", "通过持续输出有价值内容建立长期信任", "耐心等待需求触发点，不要强行催促", "一旦发现家长有需求苗头，立刻升温跟进"],
        },
    }

    tpl = templates.get(profile, templates["佛系型"])
    return {
        "script_templates": tpl["scripts"],
        "follow_up_frequency": tpl["follow_up_frequency"],
        "closing_timing": tpl["closing_timing"],
        "communication_tips": tpl["tips"],
    }


def rule_based_analysis(data: dict) -> dict:
    score = _calc_score(data)
    profile_type, profile_desc = _determine_profile(data)

    need_map = {
        "提分": "学科成绩快速提升", "升学": "升学规划与备考",
        "培养兴趣": "激发学习兴趣", "补差": "基础查漏补缺",
        "冲刺名校": "名校冲刺规划", "作业辅导": "日常作业辅助",
    }
    core_needs = [need_map[c] for c in data["concerns"] if c in need_map] or ["综合学业提升"]

    if score >= 78:   closing = "3-5天（意向强烈，及时跟进）"
    elif score >= 62: closing = "1-2周（持续培育，稳步推进）"
    elif score >= 45: closing = "2-4周（需建立信任，耐心引导）"
    else:             closing = "1个月以上（需激发需求，长期培育）"

    risks = []
    if "已在其他机构上课" in data["intent_signals"]:
        risks.append("正在竞品机构就读，转化难度较高，需突出差异化")
    if "对比竞品" in data["intent_signals"]:
        risks.append("正在多方比较，需提供清晰的差异化价值")
    if "犹豫不决" in data["intent_signals"]:
        risks.append("决策周期偏长，存在流失风险，需持续培育")
    if data["budget"] <= 3000:
        risks.append("预算偏低，报价时注意灵活处理，可先推体验班")
    if not risks:
        risks.append("暂无明显风险，保持正常节奏持续跟进即可")

    opps = []
    if "主动咨询" in data["intent_signals"]:
        opps.append("主动来询，意向明确，及时响应可快速推进")
    if "要求试听" in data["intent_signals"]:
        opps.append("已主动要求试听，转化关键节点，体验要做好")
    if data["performance"] in ("中等偏下", "较差"):
        opps.append("孩子成绩亟需提升，需求紧迫，决策动力强")
    if data["grade"] in ("初三", "高三", "初二", "高二"):
        opps.append("升学压力节点临近，时间紧迫感自然形成")
    if data["budget"] >= 15000:
        opps.append("预算充足，不存在价格障碍")
    if not opps:
        opps.append("可通过高质量内容和成功案例逐步建立信任")

    return {
        "profile_type": profile_type,
        "profile_description": profile_desc,
        "core_needs": core_needs,
        "purchase_intent_score": score,
        "closing_cycle": closing,
        "follow_up_strategy": _get_strategy(profile_type, data),
        "risk_factors": risks,
        "opportunity_points": opps,
        "analysis_method": "规则分析（离线模式）",
    }


# ─── Gemini Analysis ──────────────────────────────────────────────────────────

def analyze_with_gemini(api_key: str, data: dict, kb_snippets: str = "") -> dict:
    try:
        from google import genai
    except ImportError:
        raise RuntimeError("缺少依赖包，请执行：pip install google-genai")

    client = genai.Client(api_key=api_key)

    subjects_str = "、".join(data["subjects"])      if data["subjects"]       else "未指定"
    concerns_str = "、".join(data["concerns"])      if data["concerns"]       else "未指定"
    signals_str  = "、".join(data["intent_signals"]) if data["intent_signals"] else "未指定"
    notes_str    = data["notes"] if data["notes"] else "无"

    kb_context = f"{kb_snippets}\n\n---\n\n" if kb_snippets else ""
    prompt = f"""{kb_context}你是一名有15年经验的K12教培行业金牌销售顾问。请深度分析以下客户信息，给出专业画像和跟进策略。

## 客户信息
- 家长称呼：{data['parent_name'] or '未填写'}
- 孩子年级：{data['grade']}
- 咨询科目：{subjects_str}
- 家庭预算：{data['budget']:,}元/学期
- 孩子成绩：{data['performance']}
- 家长关注点：{concerns_str}
- 意向信号：{signals_str}
- 沟通记录：{notes_str}

## 输出要求
返回如下JSON，所有字段使用中文，话术要具体、自然、可直接使用：

{{
  "profile_type": "画像类型（焦虑型/理性分析型/价格敏感型/名校执念型/佛系型/素质培养型等，可自定义）",
  "profile_description": "约120字的详细画像描述，说明该家长心理特征、决策模式和核心驱动",
  "core_needs": ["核心诉求1（15字内）", "核心诉求2", "核心诉求3"],
  "purchase_intent_score": 0到100的整数,
  "closing_cycle": "预计成交周期（如：3-5天内 / 1-2周 / 需1个月以上）",
  "follow_up_strategy": {{
    "script_templates": [
      "话术一：完整话术内容，60-90字，语气自然，可直接说给家长听",
      "话术二：完整话术内容",
      "话术三：完整话术内容"
    ],
    "follow_up_frequency": "具体跟进频率建议（几天一次、用什么方式）",
    "closing_timing": "最佳促单时机的具体场景描述",
    "communication_tips": ["注意事项1", "注意事项2", "注意事项3", "注意事项4"]
  }},
  "risk_factors": ["风险点1（20字内）", "风险点2"],
  "opportunity_points": ["机会点1（20字内）", "机会点2", "机会点3"]
}}

只输出JSON，不要有任何其他文字或markdown标记。"""

    response = client.models.generate_content(model="gemini-2.5-flash", contents=prompt)
    text = response.text.strip()

    text = re.sub(r"^```(?:json)?\s*\n?", "", text)
    text = re.sub(r"\n?```\s*$", "", text)
    text = text.strip()

    result = json.loads(text)
    result["analysis_method"] = (
        "AI智能分析（Gemini 2.5 Flash · RAG知识库）" if kb_snippets
        else "AI智能分析（Gemini 2.5 Flash）"
    )
    return result


# ─── Render Helpers ───────────────────────────────────────────────────────────

def _score_theme(score: int):
    if score >= 80: return "#059669", "#ecfdf5", "高意向 🔥"
    if score >= 65: return "#0284c7", "#eff6ff", "中高意向"
    if score >= 45: return "#d97706", "#fffbeb", "中等意向"
    if score >= 25: return "#dc2626", "#fef2f2", "低意向"
    return "#6b7280", "#f9fafb", "意向冷淡"


def render_score_card(score: int) -> str:
    color, bg, label = _score_theme(score)
    return f"""
<div style="background:{bg}; border-radius:14px; padding:24px 16px;
            text-align:center; border:1px solid {color}28; height:100%;">
  <div style="font-size:.78rem; color:#6b7280; margin-bottom:4px; font-weight:500;">购买意向评分</div>
  <div style="font-size:3.4rem; font-weight:800; color:{color}; line-height:1.1;">{score}</div>
  <div style="background:#e5e7eb; border-radius:99px; height:8px; margin:10px 0; overflow:hidden;">
    <div style="width:{score}%; height:100%; background:{color}; border-radius:99px;"></div>
  </div>
  <div style="font-size:.9rem; font-weight:600; color:{color};">{label}</div>
</div>"""


def tags_html(items: list, cls: str = "tag-blue") -> str:
    return "".join(
        f'<span class="tag {cls}">{i}</span>' for i in items
    )


# ─── KB References Panel ──────────────────────────────────────────────────────

def render_kb_references(kb_refs: dict) -> None:
    if not kb_refs:
        return

    spin_focus = kb_refs.get("spin_focus", {})
    principles = kb_refs.get("influence_principles", [])
    examples   = kb_refs.get("spin_question_examples", [])

    st.markdown('<div class="result-section">策略依据</div>', unsafe_allow_html=True)
    st.markdown('<div class="card">', unsafe_allow_html=True)

    c1, c2 = st.columns(2)

    with c1:
        emphasis_map = {
            "多用": ("#4f46e5", "700"), "深入": ("#4f46e5", "700"),
            "慎用": ("#dc2626", "600"), "少":   ("#9ca3af", "400"),
            "轻触": ("#f59e0b", "600"), "适中": ("#64748b", "400"),
        }
        spin_rows = ""
        for key, label in [("S", "情境"), ("P", "问题发现"), ("I", "暗示后果"), ("N", "价值确认")]:
            val = spin_focus.get(key, "适中")
            color, weight = emphasis_map.get(val, ("#64748b", "400"))
            spin_rows += (
                f'<div style="display:flex;justify-content:space-between;align-items:center;'
                f'padding:6px 0;border-bottom:1px solid #f1f5f9;">'
                f'<span style="font-size:.84rem;color:#475569;"><b>{key}</b>&nbsp;{label}问题</span>'
                f'<span style="font-size:.82rem;font-weight:{weight};color:{color};">{val}</span>'
                f'</div>'
            )
        note = spin_focus.get("note", "")
        note_html = (
            f'<div style="font-size:.78rem;color:#6b7280;margin-top:8px;'
            f'line-height:1.5;font-style:italic;">{note}</div>'
        ) if note else ""
        st.markdown(f"""
        <div style="background:#f8f9ff;border-radius:10px;padding:14px 16px;">
          <div style="font-weight:700;color:#4f46e5;margin-bottom:10px;font-size:.88rem;">
            📊 SPIN 提问策略
          </div>
          {spin_rows}
          {note_html}
        </div>""", unsafe_allow_html=True)

    with c2:
        colors = ["#0ea5e9", "#8b5cf6", "#f59e0b", "#10b981"]
        cards = ""
        for i, p in enumerate(principles[:3]):
            c = colors[i % len(colors)]
            acts = " · ".join(p.get("actions", [])[:2])
            cards += (
                f'<div style="margin-bottom:10px;padding:9px 12px;background:white;'
                f'border-radius:8px;border-left:3px solid {c};">'
                f'<div style="font-weight:600;color:{c};font-size:.84rem;">{p["name"]}</div>'
                f'<div style="font-size:.78rem;color:#64748b;margin-top:3px;line-height:1.4;">'
                f'{acts}</div>'
                f'</div>'
            )
        st.markdown(f"""
        <div style="background:#f8f9ff;border-radius:10px;padding:14px 16px;">
          <div style="font-weight:700;color:#7c3aed;margin-bottom:10px;font-size:.88rem;">
            🎯 影响力原则
          </div>
          {cards or '<div style="color:#9ca3af;font-size:.85rem;">—</div>'}
        </div>""", unsafe_allow_html=True)

    if examples:
        st.markdown("<br>", unsafe_allow_html=True)
        q_rows = "".join(
            f'<div style="padding:7px 0;border-bottom:1px solid #fef9c3;'
            f'font-size:.85rem;color:#334155;">❓ {q}</div>'
            for q in examples
        )
        st.markdown(f"""
        <div style="background:#fffbeb;border-radius:10px;padding:14px 16px;
                    border:1px solid #fde68a;margin-top:2px;">
          <div style="font-weight:700;color:#92400e;margin-bottom:8px;font-size:.88rem;">
            💡 参考提问示例（I 暗示后果 / N 价值确认）
          </div>
          {q_rows}
        </div>""", unsafe_allow_html=True)

    st.markdown("</div>", unsafe_allow_html=True)


# ─── Main App ─────────────────────────────────────────────────────────────────

def main():
    config   = load_config()
    saved_key = config.get("api_key", "")

    # ── Sidebar ──────────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("""
        <div style="text-align:center; padding:10px 0 20px;">
          <div style="font-size:2.6rem;">🎓</div>
          <div style="font-weight:700; font-size:1rem; color:#1e293b; margin-top:6px;">K12客户画像系统</div>
          <div style="font-size:.78rem; color:#94a3b8; margin-top:3px;">智能销售辅助平台</div>
        </div>
        """, unsafe_allow_html=True)
        st.divider()

        st.markdown("**🔑 Gemini API 配置**")

        editing = st.session_state.get("editing_key", False)

        if saved_key and not editing:
            # Show "已配置" status
            st.markdown("""
            <div style="background:#f0fdf4;border-radius:8px;padding:8px 12px;
                        font-size:.82rem;color:#16a34a;margin-bottom:8px;">
              ✅ 已配置 · 自动读取
            </div>""", unsafe_allow_html=True)
            col_e, col_c = st.columns(2)
            with col_e:
                if st.button("修改 Key", key="btn_edit_key"):
                    st.session_state["editing_key"] = True
                    st.rerun()
            with col_c:
                if st.button("清除 Key", key="btn_clear_key"):
                    cfg = load_config()
                    cfg.pop("api_key", None)
                    save_config(cfg)
                    st.session_state.pop("editing_key", None)
                    st.rerun()
            api_key = saved_key
        else:
            api_key = st.text_input(
                "API Key", type="password",
                placeholder="粘贴 Gemini API Key...",
                help="在 Google AI Studio 免费申请",
                label_visibility="collapsed",
            )
            if editing:
                if st.button("取消", key="btn_cancel_edit"):
                    st.session_state["editing_key"] = False
                    st.rerun()
            if api_key:
                st.markdown("""
                <div style="background:#f0fdf4;border-radius:8px;padding:8px 12px;
                            font-size:.82rem;color:#16a34a;margin-top:4px;">
                  ✅ API Key 已填写，将使用 AI 分析
                </div>""", unsafe_allow_html=True)
            else:
                st.markdown("""
                <div style="background:#fffbeb;border-radius:8px;padding:8px 12px;
                            font-size:.82rem;color:#92400e;margin-top:4px;">
                  💡 未配置则使用规则分析（断网可用）
                </div>""", unsafe_allow_html=True)

        st.divider()

        st.markdown("**📚 知识库模式**")
        kb_mode = st.radio(
            "kb_mode_label",
            options=["自动（有网用RAG，断网用规则）", "仅规则", "仅RAG"],
            key="kb_mode_sel",
            label_visibility="collapsed",
            help="自动：有API时注入知识库增强AI分析；仅规则：始终离线规则；仅RAG：强制RAG（需API Key）",
        )
        st.divider()

        st.markdown("""
        <div style="font-size:.82rem; color:#64748b; line-height:1.9;">
          <b>📌 使用步骤</b><br>
          ① 填写家长和孩子信息<br>
          ② 勾选关注点和意向信号<br>
          ③ 点击「开始分析」按钮<br>
          ④ 查看画像结果与话术建议
        </div>""", unsafe_allow_html=True)
        st.divider()
        st.markdown("""
        <div style="font-size:.74rem; color:#cbd5e1; text-align:center; line-height:1.9;">
          分析结果仅供参考，数据不上传<br>
          Powered by Gemini 2.5 Flash
        </div>""", unsafe_allow_html=True)

    # ── Header ───────────────────────────────────────────────────────────────
    st.markdown("""
    <div class="hero">
      <h1>🎓 K12教培客户画像分析器</h1>
      <p>输入客户信息 · AI深度洞察 · 精准话术策略 · 提升成单率</p>
    </div>""", unsafe_allow_html=True)

    # ── Input Form ───────────────────────────────────────────────────────────
    with st.form("form", clear_on_submit=False):

        st.markdown('<div class="card"><div class="card-title">📋 基本信息</div>', unsafe_allow_html=True)
        c1, c2, c3 = st.columns(3)
        with c1:
            parent_name = st.text_input("家长称呼", placeholder="如：王女士 / 李爸爸")
        with c2:
            grade = st.selectbox("孩子年级 *", GRADE_OPTIONS)
        with c3:
            performance = st.selectbox("当前成绩水平 *", PERFORMANCE_OPTIONS)

        c4, c5 = st.columns(2)
        with c4:
            subjects = st.multiselect("咨询科目", SUBJECT_OPTIONS, help="可多选")
        with c5:
            budget = st.slider("家庭预算（元/学期）", 1000, 30000, 8000, 500)
            st.caption(f"当前预算：**¥ {budget:,}** 元/学期")
        st.markdown("</div>", unsafe_allow_html=True)

        st.markdown('<div class="card"><div class="card-title">🎯 关注点与意向信号</div>', unsafe_allow_html=True)
        c6, c7 = st.columns(2)
        with c6:
            concerns = st.multiselect("家长关注点", CONCERN_OPTIONS, help="可多选，越多越准确")
        with c7:
            intent_signals = st.multiselect("意向信号", INTENT_OPTIONS, help="根据沟通实际情况勾选")
        notes = st.text_area(
            "沟通记录 / 备注",
            placeholder="填写与家长的沟通要点、特殊需求、家庭背景等，越详细 AI 分析越精准...",
            height=88,
        )
        st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        submitted = st.form_submit_button("🔍  开始分析客户画像", use_container_width=True)

    # ── Analysis ─────────────────────────────────────────────────────────────
    if submitted:
        data = {
            "parent_name":    parent_name.strip() if parent_name else "",
            "grade":          grade,
            "subjects":       subjects,
            "budget":         budget,
            "performance":    performance,
            "concerns":       concerns,
            "intent_signals": intent_signals,
            "notes":          notes.strip() if notes else "",
        }
        display_name = data["parent_name"] or "该客户"

        with st.spinner(f"🤔 正在深度分析 {display_name} 的客户画像…"):
            result = None
            use_gemini = bool(api_key) and kb_mode != "仅规则"
            use_rag    = kb_mode in ("自动（有网用RAG，断网用规则）", "仅RAG")

            if use_gemini:
                profile_hint, _ = _determine_profile(data)
                kb_text = kb.get_rag_snippets(profile_hint, data) if use_rag else ""
                try:
                    result = analyze_with_gemini(api_key, data, kb_text)
                except Exception as e:
                    err = str(e)
                    if any(k in err.upper() for k in ("API_KEY", "INVALID", "AUTH", "PERMISSION")):
                        st.error("❌ API Key 无效，请检查后重试")
                    else:
                        st.warning(f"⚠️ AI 分析遇到问题（{err[:70]}），已自动切换至规则分析模式")
                    result = rule_based_analysis(data)
            else:
                if kb_mode == "仅RAG" and not api_key:
                    st.warning("⚠️ 仅RAG模式需要配置 Gemini API Key，已自动切换至规则分析模式")
                result = rule_based_analysis(data)

        # 附加知识库引用（供策略依据面板显示）
        result["kb_references"] = kb.get_kb_references(
            result.get("profile_type", "佛系型"), data
        )

        # 自动保存 API Key
        if api_key and api_key != saved_key:
            cfg = load_config()
            cfg["api_key"] = api_key
            save_config(cfg)
            st.session_state.pop("editing_key", None)

        # 追加历史记录
        append_to_history(data, result)

        # 写入 session_state（供下载按钮跨 rerun 使用）
        st.session_state["last_result"] = result
        st.session_state["last_data"]   = data

    # 从 session_state 取结果（下载按钮点击后 rerun 也能正常显示）
    result = st.session_state.get("last_result")
    data   = st.session_state.get("last_data")
    if result is None:
        return

    # ── Results ──────────────────────────────────────────────────────────────
    method    = result.get("analysis_method", "")
    is_ai     = "AI" in method
    m_color   = "#4f46e5" if is_ai else "#d97706"
    m_icon    = "✨" if is_ai else "📐"

    st.markdown(f"""
    <div style="display:inline-flex;align-items:center;gap:6px;background:{m_color}14;
                color:{m_color};border:1px solid {m_color}30;padding:5px 14px;
                border-radius:99px;font-size:.82rem;font-weight:600;margin:16px 0 4px;">
      {m_icon} {method}
    </div>""", unsafe_allow_html=True)

    st.markdown('<div class="result-section">画像分析</div>', unsafe_allow_html=True)

    # Score | Profile | Needs+Cycle
    score = result.get("purchase_intent_score", 50)
    cs, cp, cn = st.columns([1, 1.65, 1.35])

    with cs:
        st.markdown(render_score_card(score), unsafe_allow_html=True)

    with cp:
        ptype = result.get("profile_type", "")
        pdesc = result.get("profile_description", "")
        st.markdown(f"""
        <div style="background:white;border-radius:14px;padding:22px;
                    box-shadow:0 2px 16px rgba(0,0,0,.055);
                    border:1px solid #f1f5f9;height:100%;">
          <div style="font-size:.76rem;color:#9ca3af;margin-bottom:8px;
                      font-weight:500;letter-spacing:.5px;">客户画像类型</div>
          <div class="profile-pill">{ptype}</div>
          <div style="font-size:.86rem;color:#475569;line-height:1.75;margin-top:6px;">
            {pdesc}
          </div>
        </div>""", unsafe_allow_html=True)

    with cn:
        needs = result.get("core_needs", [])
        cycle = result.get("closing_cycle", "待评估")
        needs_t = tags_html(needs, "tag-blue")
        st.markdown(f"""
        <div style="background:white;border-radius:14px;padding:22px;
                    box-shadow:0 2px 16px rgba(0,0,0,.055);
                    border:1px solid #f1f5f9;height:100%;">
          <div style="font-size:.76rem;color:#9ca3af;margin-bottom:8px;
                      font-weight:500;letter-spacing:.5px;">核心诉求</div>
          <div style="margin-bottom:18px;min-height:48px;">{needs_t}</div>
          <div style="font-size:.76rem;color:#9ca3af;margin-bottom:5px;
                      font-weight:500;letter-spacing:.5px;">预计成交周期</div>
          <div style="font-size:.95rem;font-weight:600;color:#1e293b;">⏱ {cycle}</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Opportunities & Risks
    co, cr = st.columns(2)
    with co:
        opps = result.get("opportunity_points", [])
        rows = "".join(
            f'<div style="display:flex;gap:8px;padding:8px 0;border-bottom:1px solid #f0fdf4;">'
            f'<span style="color:#16a34a;flex-shrink:0;">✅</span>'
            f'<span style="font-size:.87rem;color:#166534;">{o}</span></div>'
            for o in opps
        )
        st.markdown(f"""
        <div style="background:white;border-radius:14px;padding:20px;
                    box-shadow:0 2px 16px rgba(0,0,0,.055);
                    border-top:3px solid #22c55e;
                    border-left:1px solid #f1f5f9;border-right:1px solid #f1f5f9;
                    border-bottom:1px solid #f1f5f9;">
          <div style="font-weight:700;color:#1e293b;margin-bottom:12px;font-size:.95rem;">
            🎯 机会点
          </div>
          {rows or '<div style="color:#9ca3af;font-size:.85rem;">暂无明显机会点</div>'}
        </div>""", unsafe_allow_html=True)

    with cr:
        risks = result.get("risk_factors", [])
        rows = "".join(
            f'<div style="display:flex;gap:8px;padding:8px 0;border-bottom:1px solid #fff7ed;">'
            f'<span style="flex-shrink:0;">⚠️</span>'
            f'<span style="font-size:.87rem;color:#9a3412;">{r}</span></div>'
            for r in risks
        )
        st.markdown(f"""
        <div style="background:white;border-radius:14px;padding:20px;
                    box-shadow:0 2px 16px rgba(0,0,0,.055);
                    border-top:3px solid #f59e0b;
                    border-left:1px solid #f1f5f9;border-right:1px solid #f1f5f9;
                    border-bottom:1px solid #f1f5f9;">
          <div style="font-weight:700;color:#1e293b;margin-bottom:12px;font-size:.95rem;">
            ⚠️ 风险点
          </div>
          {rows or '<div style="color:#9ca3af;font-size:.85rem;">暂无明显风险点</div>'}
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Follow-up Strategy
    st.markdown('<div class="result-section">跟进策略</div>', unsafe_allow_html=True)
    strategy = result.get("follow_up_strategy", {})

    st.markdown('<div class="card">', unsafe_allow_html=True)

    cf, ct = st.columns(2)
    with cf:
        freq = strategy.get("follow_up_frequency", "每周跟进一次")
        st.markdown(f"""
        <div class="info-box ib-blue">
          <div style="font-weight:600;margin-bottom:5px;">📅 跟进频率</div>
          {freq}
        </div>""", unsafe_allow_html=True)
    with ct:
        timing = strategy.get("closing_timing", "根据情况灵活把握")
        st.markdown(f"""
        <div class="info-box ib-purple">
          <div style="font-weight:600;margin-bottom:5px;">⚡ 促单时机</div>
          {timing}
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Script templates
    st.markdown('<div style="font-weight:700;color:#1e293b;margin-bottom:10px;font-size:.95rem;">💬 推荐话术模板</div>', unsafe_allow_html=True)
    for i, script in enumerate(strategy.get("script_templates", []), 1):
        st.markdown(f"""
        <div class="script-block">
          <div style="font-size:.73rem;color:#6366f1;font-weight:700;margin-bottom:6px;
                      letter-spacing:.5px;">话术模板 {i}</div>
          {script}
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Communication tips
    st.markdown('<div style="font-weight:700;color:#1e293b;margin-bottom:10px;font-size:.95rem;">💡 沟通注意事项</div>', unsafe_allow_html=True)
    tips = strategy.get("communication_tips", [])
    rows = "".join(
        f'<div class="tip-row">'
        f'<div class="tip-num">{i}</div>'
        f'<div style="font-size:.87rem;color:#475569;line-height:1.65;">{t}</div>'
        f'</div>'
        for i, t in enumerate(tips, 1)
    )
    st.markdown(f'<div style="background:#fafafa;border-radius:10px;padding:6px 14px;">{rows}</div>', unsafe_allow_html=True)

    st.markdown("</div>", unsafe_allow_html=True)

    # ── KB References ────────────────────────────────────────────────────────
    kb_refs = result.get("kb_references")
    if kb_refs:
        st.markdown("<br>", unsafe_allow_html=True)
        render_kb_references(kb_refs)

    # ── Export ───────────────────────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    report_bytes = generate_report_bytes(data, result)
    col_dl, col_info = st.columns([1, 2])
    with col_dl:
        cname    = data.get("parent_name") or "客户"
        filename = f"分析报告_{cname}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        if report_bytes:
            st.download_button(
                label="📥 导出本次报告（Excel）",
                data=report_bytes,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.warning("导出功能需要安装 openpyxl：`pip install openpyxl`")
    with col_info:
        history_tip = "（同时已自动追加至 analysis_history.xlsx）" if report_bytes else ""
        st.markdown(f"""
        <div style="font-size:.82rem;color:#94a3b8;padding-top:10px;line-height:1.7;">
          下载单个客户的完整分析报告{history_tip}
        </div>""", unsafe_allow_html=True)

    st.markdown("""
    <div style="text-align:center;padding:20px 0 6px;font-size:.77rem;color:#94a3b8;">
      分析结果仅供销售参考，请结合实际情况灵活运用 · API Key 仅保存在本地 config.json
    </div>""", unsafe_allow_html=True)


if __name__ == "__main__":
    main()
